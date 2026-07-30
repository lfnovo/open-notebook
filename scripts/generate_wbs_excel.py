"""Generate commercialization WBS / task schedule Excel workbook."""

from __future__ import annotations

from collections import OrderedDict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from wbs_tasks import TASKS  # noqa: E402

OUT = (
    Path(__file__).resolve().parents[1]
    / "docs"
    / "commercialization"
    / "Open-Notebook-Commercialization-WBS-Task-Schedule.xlsx"
)

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="1F2937")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Calibri", bold=True, size=16, color="111827")
section_font = Font(name="Calibri", bold=True, size=12, color="111827")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")

fills = {
    "Done": PatternFill("solid", fgColor="D1FAE5"),
    "In Progress": PatternFill("solid", fgColor="FEF3C7"),
    "Pending": PatternFill("solid", fgColor="E0E7FF"),
    "Blocked": PatternFill("solid", fgColor="FEE2E2"),
}
priority_fills = {
    "Critical": PatternFill("solid", fgColor="FECACA"),
    "High": PatternFill("solid", fgColor="FDE68A"),
    "Medium": PatternFill("solid", fgColor="BFDBFE"),
    "Low": PatternFill("solid", fgColor="E5E7EB"),
}

COLS = [
    ("A", "WBS ID", 10),
    ("B", "Phase / Work Package", 28),
    ("C", "Task Name", 42),
    ("D", "Detailed Description / Steps", 70),
    ("E", "Deliverable(s)", 36),
    ("F", "Status", 14),
    ("G", "Priority", 12),
    ("H", "Depends On (WBS)", 16),
    ("I", "Est. Effort (hrs)", 14),
    ("J", "Est. Duration (days)", 14),
    ("K", "Actual Effort (hrs)", 14),
    ("L", "% Complete", 12),
    ("M", "Owner / Role", 18),
    ("N", "Target Start", 12),
    ("O", "Target Finish", 12),
    ("P", "Actual Finish", 12),
    ("Q", "Acceptance Criteria", 50),
    ("R", "Risks / Notes", 40),
    ("S", "Evidence / PR / Doc", 36),
]


def style_header_row(ws, row: int = 5) -> None:
    for col, name, width in COLS:
        cell = ws[f"{col}{row}"]
        cell.value = name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 30


def main() -> None:
    tasks = TASKS
    wb = Workbook()

    ws = wb.active
    ws.title = "WBS Task List"
    ws.merge_cells("A1:S1")
    ws["A1"] = (
        "Open Notebook Commercialization — Work Breakdown Structure (WBS) & Task Schedule"
    )
    ws["A1"].font = title_font
    ws.merge_cells("A2:S2")
    ws["A2"] = (
        f"Product: DataFabricX commercial fork of lfnovo/open-notebook | "
        f"Repo: HariHaranDFX/open-notebook | Generated: {date(2026, 7, 30).isoformat()} | "
        f"Baseline: upstream-base @ v1.14.0 | Tenancy: Model A (single-tenant per client) | "
        f"Effort basis: engineering estimates (hrs); Actuals filled where known from delivered PRs"
    )
    ws["A2"].alignment = wrap
    ws.row_dimensions[2].height = 40
    ws.merge_cells("A3:S3")
    ws["A3"] = (
        "Status: Done | In Progress | Pending | Blocked | "
        "Duration assumes ~6 productive eng hrs/day | "
        "Critical path: WP0→WP1→WP-DEC→WP2→(WP3∥WP5)→WP7→WP8; WP4∥WP6 can parallel after foundations"
    )
    ws["A3"].font = Font(italic=True, size=10, color="4B5563")

    style_header_row(ws, 5)
    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:S{5 + len(tasks)}"

    for i, t in enumerate(tasks, start=6):
        for col_idx, val in enumerate(t, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin
            cell.alignment = wrap
            if col_idx in (6, 7, 9, 10, 11, 12):
                cell.alignment = center
            if col_idx == 6 and val in fills:
                cell.fill = fills[val]
                cell.font = Font(bold=True)
            if col_idx == 7 and val in priority_fills:
                cell.fill = priority_fills[val]
            if col_idx == 12 and isinstance(val, (int, float)):
                cell.value = val / 100.0
                cell.number_format = "0%"
        ws.row_dimensions[i].height = 75

    sum_ws = wb.create_sheet("Executive Summary", 0)
    sum_ws["A1"] = "Executive Summary — Commercialization Program Status"
    sum_ws["A1"].font = title_font
    sum_ws.merge_cells("A1:G1")

    headers = [
        "Work Package",
        "Status",
        "Est. Effort (hrs)",
        "Actual Effort (hrs)",
        "Est. Duration (days)",
        "% Complete (effort-weighted)",
        "Gate / Notes",
    ]
    for col, h in enumerate(headers, 1):
        c = sum_ws.cell(row=3, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = thin

    groups: OrderedDict[str, list] = OrderedDict()
    for t in tasks:
        groups.setdefault(t[1], []).append(t)

    notes = {
        "Phase 0 — R&D Diligence": "Complete before coding WPs",
        "WP0 — Foundations": "Merged PRs #1–#2",
        "WP1 — Licensing & Compliance": "Merged PRs #3–#5",
        "WP-DEC — Multi-tenancy Decision": "Model A accepted 2026-07-21",
        "Carry-forward / Tech Debt": "Not blocking WP2 start",
        "Carry-forward / Legal Follow-through": "Launch gates; not eng blockers for WP2",
        "WP2 — Identity & Entra Auth": "CURRENT — kickoff in progress",
        "WP3 — Frontend Map + White-label": "After WP2 (admin UI optional)",
        "WP4 — Backend/API Documentation": "Can parallel after WP0; better after WP2",
        "WP5 — External Source Connectors": "Hard-blocked on WP2 OAuth users",
        "WP6 — Doc Processing Benchmarks": "Needs target hardware",
        "WP7 — Cross-platform Packaging": "After WP2+WP3",
        "WP8 — Onboarding & In-app Help": "LAST — after final flows",
    }

    row = 4
    for wp, items in groups.items():
        est = sum(x[8] or 0 for x in items)
        act_vals = [x[10] for x in items if x[10] is not None]
        act = sum(act_vals) if act_vals else None
        dur = sum(x[9] or 0 for x in items)
        weighted = (
            sum((x[8] or 0) * ((x[11] or 0) / 100.0) for x in items) / est if est else 0
        )

        if wp.startswith(("WP0", "WP1", "WP-DEC", "Phase 0")):
            st = "Done"
            weighted = 1.0
        elif wp.startswith("WP2"):
            st = "In Progress"
        elif wp.startswith("Carry"):
            st = "Open follow-ups"
        else:
            st = "Pending"

        vals = [
            wp,
            st,
            est,
            act if act is not None else "—",
            round(dur, 1),
            weighted,
            notes.get(wp, ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = sum_ws.cell(row=row, column=col, value=val)
            cell.border = thin
            cell.alignment = wrap
            if col == 2:
                if st == "Done":
                    cell.fill = fills["Done"]
                elif st == "In Progress" or "Open" in st:
                    cell.fill = fills["In Progress"]
                else:
                    cell.fill = fills["Pending"]
            if col == 6 and isinstance(val, float):
                cell.number_format = "0%"
        row += 1

    for letter, width in zip("ABCDEFG", [36, 18, 16, 16, 16, 22, 48]):
        sum_ws.column_dimensions[letter].width = width

    row += 1
    total_est = sum(t[8] or 0 for t in tasks)
    total_act = sum(t[10] or 0 for t in tasks if t[10] is not None)
    total_done_est = sum(
        (t[8] or 0) * ((t[11] or 0) / 100.0) for t in tasks
    )
    sum_ws.cell(row=row, column=1, value="PROGRAM TOTAL").font = Font(bold=True)
    sum_ws.cell(row=row, column=3, value=total_est).font = Font(bold=True)
    sum_ws.cell(row=row, column=4, value=total_act).font = Font(bold=True)
    cell = sum_ws.cell(
        row=row, column=6, value=(total_done_est / total_est if total_est else 0)
    )
    cell.number_format = "0%"
    cell.font = Font(bold=True)

    row += 2
    sum_ws.cell(row=row, column=1, value="Key metrics").font = section_font
    row += 1
    metrics = [
        ("Tasks total", len(tasks)),
        ("Tasks Done", sum(1 for t in tasks if t[5] == "Done")),
        ("Tasks In Progress", sum(1 for t in tasks if t[5] == "In Progress")),
        ("Tasks Pending", sum(1 for t in tasks if t[5] == "Pending")),
        ("Tasks Blocked", sum(1 for t in tasks if t[5] == "Blocked")),
        (
            "Estimated remaining effort (hrs)",
            round(sum((t[8] or 0) * (1 - (t[11] or 0) / 100.0) for t in tasks), 1),
        ),
        (
            "Estimated remaining duration (eng-days @6h)",
            round(sum((t[9] or 0) * (1 - (t[11] or 0) / 100.0) for t in tasks), 1),
        ),
        (
            "Calendar hint (1 eng, critical path WP2→8)",
            "~10–12 weeks after WP2 start (parallelize WP4/WP6)",
        ),
        ("PRs merged (commercial fork)", "#1–#6"),
        ("Current package", "WP2 — Identity & Microsoft Entra ID (kickoff in progress)"),
    ]
    for label, val in metrics:
        sum_ws.cell(row=row, column=1, value=label).border = thin
        sum_ws.cell(row=row, column=2, value=val).border = thin
        row += 1

    row += 2
    sum_ws.cell(row=row, column=1, value="Dependency graph (critical path)").font = (
        section_font
    )
    row += 1
    sum_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    sum_ws.cell(
        row=row,
        column=1,
        value=(
            "WP0 → WP1 → WP-DEC(Model A) → WP2(Auth) → WP3(White-label) & WP5(Connectors) "
            "→ WP7(Packaging) → WP8(Onboarding)\n"
            "Parallel after foundations: WP4 (Backend map), WP6 (Benchmarks, needs hardware)\n"
            "Carry-forward tech debt / legal ToS can run alongside WP2+"
        ),
    )
    sum_ws.cell(row=row, column=1).alignment = wrap
    sum_ws.row_dimensions[row].height = 60

    ac = wb.create_sheet("Acceptance Criteria")
    ac["A1"] = "Work Package Acceptance Criteria Tracker"
    ac["A1"].font = title_font
    ac.merge_cells("A1:E1")
    for i, h in enumerate(["WP", "Criterion", "Status", "Evidence", "Notes"], 1):
        cell = ac.cell(row=3, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin

    acs = [
        ("WP0", "Clean build reproducible from docs/DEV_SETUP.md", "Done", "docs/DEV_SETUP.md", ""),
        ("WP0", "upstream-base tag exists; upstream remote configured", "Done", "git tag upstream-base", "v1.14.0 / 30c7e2a"),
        ("WP0", "Characterization tests: ingestion, search, chat, notebook CRUD", "Done", "tests/characterization/", ""),
        ("WP0", "CI runs on PR and blocks merge on failure", "Done", "7 required checks", ""),
        ("WP0", "License scan runs in CI and passes", "Done", "scripts/check_licenses.py", ""),
        ("WP1", "LICENSE retains upstream copyright + our copyright", "Done", "LICENSE", "DataFabricX Pvt Ltd"),
        ("WP1", "THIRD-PARTY-NOTICES.md complete and reproducible", "Done", "THIRD-PARTY-NOTICES.md", "Linux generate only"),
        ("WP1", "SurrealDB BSL entry present and accurate", "Done", "Notices + single image licenses/", ""),
        ("WP1", "pycountry unmodified; CI asserts", "Done", "tests/test_pycountry_unmodified.py", ""),
        ("WP1", "docs/PROVIDER_TERMS.md exists", "Done", "docs/PROVIDER_TERMS.md", "22 providers"),
        ("WP1", "No secrets committed; encryption key handling documented", "Done", "LICENSE_COMPLIANCE.md", ""),
        ("WP-DEC", "Tenancy model chosen and documented", "Done", "docs/TENANCY.md", "Model A"),
        ("WP2", "Login via real Entra account E2E", "Pending", "", "In progress package"),
        ("WP2", "Every API request validates Entra JWT (iss/aud/sig/exp)", "Pending", "", ""),
        ("WP2", "Users JIT-provisioned; user record after first login", "Pending", "", ""),
        ("WP2", "Roles enforced on sensitive routes (403)", "Pending", "", ""),
        ("WP2", "Users only see own notebooks/sources", "Pending", "", ""),
        ("WP2", "Password provider still works as local-dev fallback", "Pending", "", ""),
        ("WP2", "Second provider needs no route-handler changes (stub)", "Pending", "", ""),
        ("WP2", "docs/AUTH.md complete", "Pending", "", ""),
        ("WP3", "FRONTEND_MAP.md complete with where-to-change table", "Pending", "", ""),
        ("WP3", "New client branded by config only (two brands, same build)", "Pending", "", ""),
        ("WP3", "Hardcoded brand strings/assets read from BrandConfig", "Pending", "", ""),
        ("WP3", "Dark/light both themed correctly", "Pending", "", ""),
        ("WP4", "Every router/endpoint documented with auth + models", "Pending", "", ""),
        ("WP4", "Every domain model and graph documented", "Pending", "", ""),
        ("WP4", "Five how-to-add-X playbooks present", "Pending", "", ""),
        ("WP4", "New engineer can locate changes using only BACKEND_MAP", "Pending", "", ""),
        ("WP5", "SharePoint connect → folder → pull all supported docs", "Pending", "", ""),
        ("WP5", "Ingestion reuses existing pipeline", "Pending", "", ""),
        ("WP5", "Batch per-doc progress; single failure non-fatal", "Pending", "", ""),
        ("WP5", "Chat across whole pulled set", "Pending", "", ""),
        ("WP5", "Tokens encrypted per user; source permissions respected", "Pending", "", ""),
        ("WP5", "New connector = SourceConnector interface only", "Pending", "", ""),
        ("WP5", "docs/CONNECTORS.md complete", "Pending", "", ""),
        ("WP6", "Benchmark harness repeatable with metrics report", "Pending", "", ""),
        ("WP6", "PERFORMANCE_AND_SIZING.md has measured numbers (2 profiles)", "Pending", "", ""),
        ("WP6", "Supported types + tested size limits + failure points", "Pending", "", ""),
        ("WP6", "Accuracy methodology + results per model tier", "Pending", "", ""),
        ("WP6", "Min vs recommended hardware justified by numbers", "Pending", "", ""),
        ("WP7", "Container deploy verified Windows + Linux", "Pending", "", ""),
        ("WP7", "Native install documented/verified Linux + Windows", "Pending", "", ""),
        ("WP7", "Installer CLI full install interactive + non-interactive", "Pending", "", ""),
        ("WP7", "Scripted install stands up branded client E2E", "Pending", "", ""),
        ("WP7", "docs/DEPLOYMENT.md complete", "Pending", "", ""),
        ("WP8", "Fresh instance launches setup wizard → working app", "Pending", "", ""),
        ("WP8", "Guided tour of core loop completable", "Pending", "", ""),
        ("WP8", "Contextual tooltips; all copy in i18n", "Pending", "", ""),
        ("WP8", "Every main screen useful empty state", "Pending", "", ""),
        ("WP8", "Tour dismissible/resumable; help surface everywhere", "Pending", "", ""),
        ("WP8", "docs/ONBOARDING.md complete", "Pending", "", ""),
        ("Program", "New client by config + installer only (no per-client rebuild)", "Pending", "", "Definition of Done"),
        ("Program", "Licensing compliant and drift-guarded", "Done", "WP1 + CI License Scan", ""),
        ("Program", "Real performance/sizing numbers for clients", "Pending", "WP6", ""),
        ("Program", "Backend + frontend fully mapped", "Pending", "WP3a + WP4", ""),
    ]
    for i, rowv in enumerate(acs, 4):
        for c, v in enumerate(rowv, 1):
            cell = ac.cell(row=i, column=c, value=v)
            cell.border = thin
            cell.alignment = wrap
            if c == 3 and v in fills:
                cell.fill = fills[v]
                cell.font = Font(bold=True)
    ac.column_dimensions["A"].width = 12
    ac.column_dimensions["B"].width = 70
    ac.column_dimensions["C"].width = 12
    ac.column_dimensions["D"].width = 36
    ac.column_dimensions["E"].width = 28
    ac.freeze_panes = "A4"

    asm = wb.create_sheet("Estimating Assumptions")
    asm["A1"] = "Estimating Assumptions & How to Use This Workbook"
    asm["A1"].font = title_font
    asm.merge_cells("A1:B1")
    asm["A3"] = "Topic"
    asm["B3"] = "Detail"
    asm["A3"].fill = header_fill
    asm["B3"].fill = header_fill
    asm["A3"].font = header_font
    asm["B3"].font = header_font
    assumptions = [
        ("Effort unit", "Hours of experienced engineer time (design + implement + test + docs + PR)."),
        ("Duration", "Working days assuming ~6 productive hours/day (meetings/review buffer already outside)."),
        ("Team model", "Default schedule assumes 1 full-time engineer; calendar time shrinks with parallel WPs (WP4/WP6 vs WP2)."),
        ("Actuals", "Actual Effort is filled for completed work from delivered PRs and recorded close-out dates."),
        ("Contingency", "No pad in task estimates. Recommend +20% management reserve on remaining program for WP2 auth/ownership surprises."),
        ("Critical path driver", "WP2 (auth + data ownership) — largest single risk and gate for WP3/WP5."),
        ("What Done means", "Acceptance criteria pass AND tests green AND PR merged to protected main AND human review stop completed."),
        ("Not in estimates", "Formal counsel fees; Entra tenant procurement; client UAT; sales/marketing; paid hardware for WP6 if unavailable."),
        ("Currency of plan", "Synced from the master implementation plan, LEGAL_DECISIONS/TENANCY docs, and git main as of 2026-07-30."),
        ("Update cadence", "Update Status / Actual Effort / % Complete at end of each WP; never mix WPs on one branch."),
        ("Sheet guide", "Executive Summary = leadership; WBS Task List = full schedule; Acceptance Criteria = DoD checklist; this sheet = estimating rules."),
        ("Regenerate", "uv run python scripts/generate_wbs_excel.py (task data lives in scripts/wbs_tasks.py)."),
    ]
    for i, (a, b) in enumerate(assumptions, 4):
        asm.cell(row=i, column=1, value=a).border = thin
        cell = asm.cell(row=i, column=2, value=b)
        cell.border = thin
        cell.alignment = wrap
        asm.row_dimensions[i].height = 35
    asm.column_dimensions["A"].width = 24
    asm.column_dimensions["B"].width = 110

    # Sanity: no coding-assistant / agent marketing language in workbook payload.
    # (Repo filenames like CLAUDE.md are fine; product terms like "AI providers" are fine.)
    banned = (
        "claude code",
        "claude session",
        "chatgpt",
        "ai agent",
        "ai-assisted",
        "ai assisted",
        "cursor agent",
        "coding agent",
    )
    blob = " ".join(str(x) for t in tasks for x in t).lower()
    for b in banned:
        if b in blob:
            raise SystemExit(f"Banned phrase found in task data: {b!r}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Tasks: {len(tasks)}")
    print(f"Total est hrs: {total_est}")
    print(f"Effort-weighted complete hrs: {round(total_done_est, 1)}")
    print(f"In Progress: {[t[0] for t in tasks if t[5] == 'In Progress']}")
    print(f"Has 3.12: {any(t[0] == '3.12' for t in tasks)}")


if __name__ == "__main__":
    main()
